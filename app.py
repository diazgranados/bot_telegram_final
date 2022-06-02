
from fileinput import filename
from telegram.ext import (Updater, CommandHandler, MessageHandler, Filters)
import mysql.connector 
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import time
from openpyxl.chart import ScatterChart, Reference, Series
from telegram.ext.callbackcontext import CallbackContext


db=mysql.connector.connect(
    host="localhost",
    user="root",
    password="",
    port=3306,
    database="bot_telegram",
)
#app = Flask(__name__)
#mysql = MySQL(app)








#update.message.text() resive el texto
def help(update, context):
	''' comandos '''
	
	context.bot.send_message(update.message.chat_id,"Bienvenido \n estos son nuestros comandos \n /ingresos_mes : este comando te permitira saber todos los ingresos que has tenido el mes actual \n /mostrar_gastos con este comando veras todos los gastos que has hecho \n /ingresar_monto : con este podras realizar el guardado de tu monto para llevarlo en el registro,recuerda ingresar el monto junto a la descripcion y al comando poner  fecha a tu monto ingresado  formato 2022-05-25 \n /ingresar_gasto exactamente igual que ingresar monto solo que en este apartado ingresaras un gasto \n /suma_ingreso te mostrara el resultado de la suma de tus ingresos  \n /suma_gasto te mostrara el resultado de la suma de tus gastos \n /excel_ingreso creara un archivo excel con los datos de tus ingresos \n /excel_gasto creara un archivo excel con los datos de tus gastos \n /grafico_ingreso te mostrara un excel con un grafico de los ingresos que has tenido")
def suma(update,context):
    	
    #context.bot.send_message(update.message.chat_id,"Hola, por favor escribenos en numeros sin espacios ni comas el monto que deseas ingresar")

	
	user = update.message.from_user.id
	cursor = db.cursor()
	cursor.execute("SELECT SUM(monto_a_ingresar)AS suma FROM ingresos where id_user=%s",(user,))
	suma=cursor.fetchone()
	db.commit()
	#update.message.reply_text("por favor utilice dos numeros",suma[0])
	context.bot.send_message(update.message.chat_id,"La sumatoria de tus ingresos es: "+str(suma[0]))

def ingresar_monto(update,context):
    	
    #context.bot.send_message(update.message.chat_id,"Hola, por favor escribenos en numeros sin espacios ni comas el monto que deseas ingresar")
	user = update.message.from_user.id
	numero1 = int(context.args[0])
	fecha = context.args[1]
	descripcion=context.args[2]

	cursor = db.cursor()
	cursor.execute("INSERT INTO ingresos (id_user,monto_a_ingresar,fecha,descripcion)VALUES (%s,%s,%s,%s)", (user,numero1,descripcion,fecha))
	db.commit()

def ingresos_mes(update, context):

	"""
	user = update.message.from_user.id
	cursor=db.cursor()
	cursor.execute("SELECT monto_a_ingresar,descripcion,fecha FROM ingresos where id_user=%s",(user,))

	mes=cursor.fetchall()
	texto=""
	for database in mes:
    	
		texto = texto + '\n'+str(database[0])+" "+str(database[1])+" "+str(database[2])
		print(database)
	context.bot.send_message(update.message.chat_id,"Hola, acontinuacion te mostraremos los ingresos de este mes  "+texto)


"""

	fecha1 = context.args[0]
	fecha2 = context.args[1]
	user = update.message.from_user.id
	cursor = db.cursor()
	cursor.execute("SELECT  monto_a_ingresar,descripcion,fecha from ingresos WHERE fecha>=%s AND fecha<=%s AND id_user=%s",(fecha1,fecha2,user,))
	ingreso= cursor.fetchall()

	

	book = Workbook()
	sheet = book.active
#	for i in database[i]:
#		sheet[f'A{i}'] = database[i]	
	sheet['B1'] = 'descripcion'

	sheet['C1'] = 'fecha'

	sheet['A1'] = "monto a ingresar"
	for i,value in enumerate(ingreso):
		sheet[f'A{i+2}'] = value[0]
		sheet[f'B{i+2}'] = value[1]
		sheet[f'C{i+2}'] = value[2]
	book.save('ingresos_mes.xlsx')		
	context.bot.send_document(update.message.chat_id,document=open("ingresos_mes.xlsx","rb"),filename="ingresos_mes.xlsx") 	

def mostrar_gastos(update,context):
	user = update.message.from_user.id
	cursor=db.cursor()
	cursor.execute("SELECT ingreso_gasto,descripcion,fecha FROM gastos where id_user=%s",(user,))

	mes=cursor.fetchall()
	texto=""
	for database in mes:
    	
		texto = texto + '\n'+str(database[0])+" "+str(database[1])+" "+str(database[2])
		print(database)
	context.bot.send_message(update.message.chat_id,"Hola, acontinuacion te mostraremos los ingresos de este mes  "+texto)    		
    	
def start(update, context: CallbackContext):
	''' START '''
	# Enviar un mensaje a un ID determinado.
	print(update.message.from_user.id)


	context.bot.send_message(update.message.chat_id, "Hola, bienvenido a tu bot de confianza, llevare un registro de tus ingresos y gastos para que puedas organizarte mejor, escribe /help para saber como funciono")

def gasto(update,context):
	numero1 = int(context.args[0])
	fecha = context.args[1]
	descripcion=context.args[2]
	user = update.message.from_user.id
	cursor = db.cursor()
	cursor.execute("INSERT INTO gastos (id_user,ingreso_gasto,fecha,descripcion)VALUES (%s,%s,%s,%s)", (user,numero1,descripcion,fecha))
	db.commit()

def suma_gasto(update,context ):
	cursor = db.cursor()
	cursor.execute("SELECT SUM(ingreso_gasto)AS GASTOS FROM gastos ")
	suma=cursor.fetchone()
	db.commit()
	#update.message.reply_text("por favor utilice dos numeros",suma[0])
	context.bot.send_message(update.message.chat_id,"La sumatoria de tus Gastos es: "+str(suma[0]))
def excel_ingreso(update,context):
	user = update.message.from_user.id
	cursor = db.cursor()
	cursor.execute("SELECT monto_a_ingresar,descripcion,fecha FROM ingresos where id_user=%s",(user,))
	ingreso= cursor.fetchall()

	

	book = Workbook()
	sheet = book.active
#	for i in database[i]:
#		sheet[f'A{i}'] = database[i]	
	sheet['B1'] = 'descripcion'

	sheet['C1'] = 'fecha'

	sheet['A1'] = "monto a ingresar"
	for i,value in enumerate(ingreso):
		sheet[f'A{i+2}'] = value[0]
		sheet[f'B{i+2}'] = value[1]
		sheet[f'C{i+2}'] = value[2]
	#sheet['A2'] = database[0]



	#sheet['B1'].font = Font(color='FF0000', bold=True)
	#for i in range(2, 15):
	#	sheet[f'B{i}'] = i**2

	sheet2 = book.create_sheet('hoja_2')
	sheet2['A1'] = 'SUSCRIBETE'
	fecha = time.strftime('%x')
	sheet2['A2'] = fecha

# sheet3.unmerge_cells('A1:D1')


	book.save('ingreso.xlsx')    	
	context.bot.send_document(update.message.chat_id,document=open("ingreso.xlsx","rb"),filename="ingreso.xlsx") 

def excel_gasto(update,context):
	user = update.message.from_user.id  	
	cursor = db.cursor()
	cursor.execute("SELECT ingreso_gasto,descripcion,fecha FROM gastos where id_user=%s",(user,))
	ingreso= cursor.fetchall()
	texto=""
	for database in ingreso:
        	
		texto =str(database[0])+" "+str(database[1])+" "+str(database[2])
		print(database)
	

	book = Workbook()
	sheet = book.active
#	for i in database[i]:
#		sheet[f'A{i}'] = database[i]	
	sheet['B1'] = 'descripcion'

	sheet['C1'] = 'fecha'

	sheet['A1'] = "monto a ingresar"
	for i,value in enumerate(ingreso):
		sheet[f'A{i+2}'] = value[0]
		sheet[f'B{i+2}'] = value[1]
		sheet[f'C{i+2}'] = value[2]	
#	context.bot.send_message(update.message.chat_id,)	
	book.save('gastos.xlsx')
	context.bot.send_document(update.message.chat_id,document=open("gastos.xlsx","rb"),filename="gastos.xlsx")    	
def grafico_ingreso(update,context):
	book = Workbook()
	sheet = book.active

	for i in range(1, 15):
		sheet[f'A{i}'] = i

	for i in range(1, 15):
		sheet[f'B{i}'] = i*10

	c1 = ScatterChart()
	c1.title = 'Gráfico de Dispersión'
	c1.style = 13
	c1.y_axis.title = 'eje Y'
	c1.x_axis.title = 'eje X'

	xvalues = Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=14)
	yvalues = Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=14)
	ser = Series(yvalues, xvalues, title='recta')
	c1.series.append(ser)

	sheet.add_chart(c1, 'D3')

	book.save('grafico_ingreso.xlsx') 	


def main():
	TOKEN="5368500412:AAFZSD-TXNe_MWOaLQbcFc6AzpqCDYxTIYQ"
	updater=Updater(TOKEN, use_context=True)
	dp=updater.dispatcher

	# Eventos que activarán nuestro bot.
	dp.add_handler(CommandHandler('start',	start))
	dp.add_handler(CommandHandler('help',	help))
	dp.add_handler(CommandHandler('ingresos_mes',	ingresos_mes))
	dp.add_handler(CommandHandler('ingresar_monto',	ingresar_monto))
	dp.add_handler(CommandHandler('ingresar_gasto',	gasto))
	dp.add_handler(CommandHandler('suma_ingreso',	suma))
	dp.add_handler(CommandHandler('suma_gasto',	suma_gasto))
	dp.add_handler(CommandHandler('excel_ingreso',	excel_ingreso))
	dp.add_handler(CommandHandler('excel_gasto',	excel_gasto))
	dp.add_handler(CommandHandler('grafico_ingreso', grafico_ingreso))
	dp.add_handler(CommandHandler('mostrar_gastos', mostrar_gastos))



	# Comienza el bot
	#dp.add_handler(MessageHandler(Filters.text, ))
	updater.start_polling()
	# Lo deja a la escucha. Evita que se detenga.
	updater.idle()

if __name__ == '__main__':
	main()


